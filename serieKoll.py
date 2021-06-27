# py
# INSTRUKTIONER: Lägg programmet i mappen med videofiler. En excelfil skapas automatiskt som håller koll på vilka 
# avsnitt som tittats på. Starta programmet för att fortsätta på rätta avsnitt. 

# Detta är en stabil build 

from pathlib import Path
import openpyxl, os, datetime, time, statistics, pyinputplus

# Folder för videos
folder = os.path.dirname(os.path.realpath(__file__)) # Få mappen där filen körs ifrån
folder = Path(folder)

# Läs in filer
filer = list(folder.glob('*')) # Lägger filnamnen i en lista

# ändra CWD 
os.chdir(folder)

# Sätt upp listor
videoLista = [] 
filändelseLista = []
allaFiltyper = []

# loopar över filer för att kolla filändelser
for filnamn in filer:
    
    filändelse = '.' + str(filnamn.name).split('.')[-1]
    filändelseLista.append(filändelse) # Lägger till alla filändelser i en lista 

vanligasteFiltyp = statistics.mode(filändelseLista) # Ta fram vanligaste filtyp
allaFiltyper = list(set(filändelseLista)) # Skapa en lista med alla filtyper

if '.xlsx' in allaFiltyper:   # Om det redan finns en excelfil
    print('Hittade excelfil!')
    wb = openpyxl.load_workbook(str(folder.name) + '.xlsx')
    print('Excelfil laddad...')
    sheet = wb.active
    filtyp = '.' + sheet.cell(column=1, row=2).value.split('.')[-1]
    print(filtyp)

else: # Om det inte finns en excelfil
    print('Det verkar som att du tittar på den här serien första gången...')
    print('Är videon av filtypen: '+ str(vanligasteFiltyp) + '?')
    svar = pyinputplus.inputMenu(['Ja', 'Nej'], numbered=True)

    if svar == 'Ja':
        filtyp = '.' + str(vanligasteFiltyp)
    else:
        print('Vilken filtyp är videofilerna av?')
        svar = pyinputplus.inputMenu(allaFiltyper, numbered=True)
        filtyp = '.' + str(svar)

    print('Skapar ny excelfil')
    wb = openpyxl.Workbook() # Skapar ny workbook

# loopar över filer
for filnamn in filer:
    if str(filnamn.name).split('.')[-1] in filtyp: # Letar efter filändelsen
        videoLista.append(str(filnamn))    #  Lägger till i en videolista # TODO Är det här jag ska plocka hela path istället?

# Starta while-loop här
while True:
    # Kolla igenom första kolumnen 
    sheet = wb.active
    tittadeVideos = []
    for row in range(1,sheet.max_row+1):  # Loopar över raderna i kolunmn 1
        tittadeVideos.append(sheet.cell(row=row, column=1).value)      # Lägger filnamnen i kolumn 1 i en lista

    if tittadeVideos != [None]: # Om listan med tittade videos inte är tom:
        for i in tittadeVideos:  # Loopa över tittade videos
            if i in videoLista: 
                print(str(i) +' har du redan sett. Jag tar fram nästa video...') # Plockar ut filnamnet från filepath 
                videoLista.remove(videoLista[0])     # Ta bort först elementet i listan med alla videos 
    else:
        print('Du har inte sett några videos än...')
        print('Startar första videon...')

    körFil = videoLista[0] # Plocka ut första elementet i videoListan som körfil.
    
    print()
    print('Körfilen är: '+ körFil)
    input('Tryck enter när du är redo...')
    # Lägg till den körda filen i excelfilen
    sheet.cell(row=sheet.max_row+1, column=1).value = str(körFil)

    # Kör igång filen. OBS! Koden väntar på att videofönstret stängs innan den fortsätter.  
    print()
    print('Kör igång ' + str(körFil.split('\\')[-1]))
    print()
    print('OBS! Stäng inte det här fönstret för då blir sparfilen korrupt.')
    os.system('"' + körFil + '"') 
    
    print('Sparar excelfil...')
    wb.save(str(folder.name) + '.xlsx')

    print()
    print('Vill du se ett avsnitt till?')
    slutMeny = ['Spela en till', 'Avsluta']
    slutVal = pyinputplus.inputMenu(slutMeny, numbered=True)
    if slutVal == 'Avsluta':
        break
    else:
        print('Ok, då kör vi en till...')

print('DONE!')

